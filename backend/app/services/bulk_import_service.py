"""BulkImportService — template generation, schema introspection, and server-side
sheet parsing for the file-based bulk import flow (Wave B).
"""
from __future__ import annotations

import csv
import io
import mimetypes
import zipfile
from typing import Any
from uuid import UUID

from fastapi import HTTPException
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models.user import User
from app.schemas.products import (
    BulkDraftCreated,
    BulkDraftImage,
    BulkDraftRow,
    BulkSaveDraftsRequest,
)
from app.services.bulk_import.columns import COLUMNS
from app.services.bulk_import.template import build_template_xlsx
from app.services.file_service import FileService
from app.utils.logging import LoggerMixin


class BulkImportService(LoggerMixin):
    def __init__(self, session: AsyncSession | None = None) -> None:
        """Session is optional because get_template_xlsx() and get_schema() are pure."""
        self.session = session

    # ------------------------------------------------------------------
    # Pure helpers (no DB)
    # ------------------------------------------------------------------

    def get_template_xlsx(self) -> bytes:
        """Returns the xlsx bytes from build_template_xlsx()."""
        self.logger.debug("get_template_xlsx called")
        return build_template_xlsx()

    def get_schema(self) -> dict:
        """Returns column metadata in the same shape as the bulk_schema endpoint.

        Shape:
            {
                "columns": [
                    {key, title, required, example, description, aliases: list},
                    ...
                ]
            }
        """
        self.logger.debug("get_schema called columns_count=%d", len(COLUMNS))
        return {
            "columns": [
                {
                    "key": c.key,
                    "title": c.title,
                    "required": c.required,
                    "example": c.example,
                    "description": c.description,
                    "aliases": list(c.aliases),
                }
                for c in COLUMNS
            ]
        }

    # ------------------------------------------------------------------
    # Sheet parsing helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _build_auto_column_mapping(sheet_headers: list[str]) -> dict[str, str]:
        """Auto-detect a {sheet_header: canonical_key} mapping.

        For each sheet header, check if it matches a ColumnSpec title or any
        alias (case-insensitive, trimmed).  Headers that match nothing are
        omitted (they will be ignored during row building).
        """
        mapping: dict[str, str] = {}
        for header in sheet_headers:
            normalised = header.strip().lower()
            for spec in COLUMNS:
                # Check canonical title
                if spec.title.strip().lower() == normalised:
                    mapping[header] = spec.key
                    break
                # Check aliases
                if any(a.strip().lower() == normalised for a in spec.aliases):
                    mapping[header] = spec.key
                    break
        return mapping

    @staticmethod
    def _parse_xlsx(raw: bytes) -> list[dict[str, str]]:
        """Parse xlsx/xls bytes into a list of {header: cell_value} dicts.

        Uses openpyxl for xlsx (and xls via compatibility mode).  Only the
        first worksheet is read.  Rows where every cell is empty are skipped.
        """
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []

        headers = [str(h).strip() if h is not None else "" for h in header_row]

        result: list[dict[str, str]] = []
        for row in rows_iter:
            cells = [str(c).strip() if c is not None else "" for c in row]
            # Skip fully empty rows
            if all(v == "" or v == "None" for v in cells):
                continue
            result.append(dict(zip(headers, cells)))

        wb.close()
        return result

    @staticmethod
    def _parse_csv(raw: bytes) -> list[dict[str, str]]:
        """Parse CSV bytes into a list of {header: cell_value} dicts.

        Tries UTF-8 first, then falls back to Windows-1256 (common for
        Persian spreadsheets exported from Excel).
        """
        for encoding in ("utf-8-sig", "utf-8", "windows-1256"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            # Last resort — ignore errors
            text = raw.decode("utf-8", errors="replace")

        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            clean = {k.strip(): v.strip() for k, v in row.items() if k}
            if any(v for v in clean.values()):
                rows.append(clean)
        return rows

    # ------------------------------------------------------------------
    # Row → BulkDraftRow conversion
    # ------------------------------------------------------------------

    @staticmethod
    def _coerce_int(value: str) -> int | None:
        """Coerce a string cell to int, returning None on failure."""
        try:
            return int(float(value.replace(",", "").strip()))
        except (ValueError, AttributeError):
            return None

    @staticmethod
    def _coerce_float(value: str) -> float | None:
        """Coerce a string cell to float, returning None on failure."""
        try:
            return float(value.replace(",", "").strip())
        except (ValueError, AttributeError):
            return None

    @staticmethod
    def _coerce_bool(value: str) -> bool | None:
        """Coerce 'true'/'false'/'1'/'0' to bool, returning None on failure."""
        v = value.strip().lower()
        if v in ("true", "1", "yes", "بله"):
            return True
        if v in ("false", "0", "no", "خیر"):
            return False
        return None

    def _build_draft_row(
        self,
        raw_row: dict[str, str],
        mapping: dict[str, str],
    ) -> dict[str, Any]:
        """Convert a raw header→cell dict into canonical key→value dict.

        Returns a plain dict with canonical keys.  Integer/bool coercion is
        applied where the ColumnSpec dictates.  Missing keys are omitted
        (caller checks required fields).
        """
        canonical: dict[str, Any] = {}
        for sheet_header, cell_value in raw_row.items():
            key = mapping.get(sheet_header)
            if key is None or cell_value == "" or cell_value == "None":
                continue

            if key in ("primary_price", "stock", "weight_g", "package_weight_g",
                       "preparation_days"):
                coerced = self._coerce_int(cell_value)
                if coerced is not None:
                    canonical[key] = coerced
            elif key == "unit_quantity":
                coerced_f = self._coerce_float(cell_value)
                if coerced_f is not None:
                    canonical[key] = coerced_f
            elif key == "is_wholesale":
                coerced_b = self._coerce_bool(cell_value)
                if coerced_b is not None:
                    canonical[key] = coerced_b
            elif key == "keywords":
                # Store as list — split by comma
                canonical[key] = [k.strip() for k in cell_value.split(",") if k.strip()]
            else:
                canonical[key] = cell_value

        return canonical

    # ------------------------------------------------------------------
    # Main entry point: save_drafts_from_files
    # ------------------------------------------------------------------

    async def save_drafts_from_files(
        self,
        user: User,
        *,
        sheet_file_id: UUID,
        zip_file_id: UUID | None,
        column_mapping: dict[str, str] | None,
    ) -> list[BulkDraftCreated]:
        """Parse a pre-uploaded sheet (xlsx/csv) and optional ZIP of images,
        then persist all rows as DRAFT products.

        Args:
            user: Authenticated user who owns the files and will own the drafts.
            sheet_file_id: UUID of a ``File`` row with ``kind="bulk_sheet"``.
            zip_file_id: UUID of a ``File`` row with ``kind="bulk_zip"``, or
                ``None`` when no ZIP was uploaded.
            column_mapping: Optional explicit ``{sheet_header: canonical_key}``
                mapping.  When ``None`` the mapping is auto-detected from the
                sheet headers using COLUMNS title/aliases.

        Returns:
            List of ``BulkDraftCreated`` (same shape as the rows-based endpoint).

        Raises:
            HTTPException 400: When a file has the wrong ``kind``, the format
                is unrecognised, or required columns are absent in the sheet.
            HTTPException 404: When a file ID is not found or not owned by the
                user (raised by ``FileService.get_for_user``).
        """
        assert self.session is not None, "session required for save_drafts_from_files"

        file_svc = FileService(self.session)

        # ------------------------------------------------------------------ #
        # 1. Resolve + validate sheet file
        # ------------------------------------------------------------------ #
        sheet_file = await file_svc.get_for_user(sheet_file_id, user)
        if sheet_file.kind != "bulk_sheet":
            self.logger.warning(
                "save_drafts_from_files wrong kind file_id=%s kind=%s user_id=%s",
                sheet_file_id, sheet_file.kind, user.id,
            )
            raise HTTPException(
                status_code=400,
                detail="فایل شیت باید با نوع bulk_sheet آپلود شده باشد.",
            )

        # ------------------------------------------------------------------ #
        # 2. Read + parse sheet
        # ------------------------------------------------------------------ #
        sheet_bytes = await file_svc.read_bytes(sheet_file)
        filename_lower = (sheet_file.filename or "").lower()

        if filename_lower.endswith(".csv"):
            raw_rows = self._parse_csv(sheet_bytes)
            self.logger.debug(
                "save_drafts_from_files parsed csv rows=%d file_id=%s",
                len(raw_rows), sheet_file_id,
            )
        elif filename_lower.endswith((".xlsx", ".xls")):
            raw_rows = self._parse_xlsx(sheet_bytes)
            self.logger.debug(
                "save_drafts_from_files parsed xlsx rows=%d file_id=%s",
                len(raw_rows), sheet_file_id,
            )
        else:
            self.logger.error(
                "save_drafts_from_files unrecognised format filename=%s file_id=%s",
                sheet_file.filename, sheet_file_id,
            )
            raise HTTPException(
                status_code=400,
                detail="فرمت فایل شیت پشتیبانی نمی‌شود. فقط xlsx، xls و csv مجاز است.",
            )

        if not raw_rows:
            raise HTTPException(
                status_code=400,
                detail="فایل شیت خالی است یا ردیف داده‌ای ندارد.",
            )

        # ------------------------------------------------------------------ #
        # 3. Resolve column mapping
        # ------------------------------------------------------------------ #
        all_headers = list(raw_rows[0].keys())

        if column_mapping is not None:
            effective_mapping = column_mapping
            self.logger.debug(
                "save_drafts_from_files using provided column_mapping headers=%s",
                list(effective_mapping.keys()),
            )
        else:
            effective_mapping = self._build_auto_column_mapping(all_headers)
            self.logger.debug(
                "save_drafts_from_files auto-detected mapping=%s", effective_mapping,
            )

        # ------------------------------------------------------------------ #
        # 4. Resolve + validate ZIP file (optional)
        # ------------------------------------------------------------------ #
        # Map normalised filename → (zip_member_name) for images/ entries.
        zip_image_map: dict[str, str] = {}
        zip_bytes_cache: bytes | None = None

        if zip_file_id is not None:
            zip_file = await file_svc.get_for_user(zip_file_id, user)
            if zip_file.kind != "bulk_zip":
                self.logger.warning(
                    "save_drafts_from_files wrong kind zip file_id=%s kind=%s user_id=%s",
                    zip_file_id, zip_file.kind, user.id,
                )
                raise HTTPException(
                    status_code=400,
                    detail="فایل ZIP باید با نوع bulk_zip آپلود شده باشد.",
                )

            zip_bytes_cache = await file_svc.read_bytes(zip_file)
            try:
                with zipfile.ZipFile(io.BytesIO(zip_bytes_cache)) as zf:
                    for member in zf.namelist():
                        # Only index entries under images/ (case-insensitive prefix)
                        if member.lower().startswith("images/"):
                            basename = member.split("/")[-1]
                            if basename:
                                zip_image_map[basename.lower()] = member
            except zipfile.BadZipFile as exc:
                self.logger.error(
                    "save_drafts_from_files bad zip file_id=%s error=%s",
                    zip_file_id, exc,
                )
                raise HTTPException(
                    status_code=400,
                    detail="فایل ZIP معتبر نیست یا خراب است.",
                )

            self.logger.debug(
                "save_drafts_from_files zip indexed entries=%d file_id=%s",
                len(zip_image_map), zip_file_id,
            )

        # ------------------------------------------------------------------ #
        # 5. Build BulkDraftRow objects, extract images from ZIP
        # ------------------------------------------------------------------ #
        required_keys = {c.key for c in COLUMNS if c.required}
        draft_rows: list[BulkDraftRow] = []

        for row_idx, raw_row in enumerate(raw_rows):
            canonical = self._build_draft_row(raw_row, effective_mapping)

            # Skip rows missing any required column
            missing = required_keys - canonical.keys()
            if missing:
                self.logger.warning(
                    "save_drafts_from_files skipping row_idx=%d missing_keys=%s",
                    row_idx, missing,
                )
                continue

            # ---- Image handling ----------------------------------------- #
            images: list[BulkDraftImage] = []

            # 5a: Images referenced by filename in the ZIP
            image_filenames_raw: str = canonical.pop("image_filenames", "") or ""
            if image_filenames_raw and zip_bytes_cache is not None:
                filenames = [f.strip() for f in image_filenames_raw.split(",") if f.strip()]
                for fname in filenames:
                    member_name = zip_image_map.get(fname.lower())
                    if member_name is None:
                        self.logger.warning(
                            "save_drafts_from_files row_idx=%d image not found in zip fname=%s",
                            row_idx, fname,
                        )
                        continue

                    try:
                        with zipfile.ZipFile(io.BytesIO(zip_bytes_cache)) as zf:
                            img_bytes = zf.read(member_name)
                    except Exception as exc:
                        self.logger.warning(
                            "save_drafts_from_files row_idx=%d zip read error fname=%s error=%s",
                            row_idx, fname, exc,
                        )
                        continue

                    mime, _ = mimetypes.guess_type(fname)
                    mime = mime or "application/octet-stream"

                    try:
                        uploaded_file = await file_svc.upload_from_bytes(
                            user=user,
                            raw=img_bytes,
                            filename=fname,
                            mime=mime,
                            kind="product_image",
                        )
                    except HTTPException as exc:
                        self.logger.warning(
                            "save_drafts_from_files row_idx=%d upload failed fname=%s status=%s detail=%s",
                            row_idx, fname, exc.status_code, exc.detail,
                        )
                        continue

                    # Prefer file_id path; fall back to data_url if field absent
                    if hasattr(BulkDraftImage, "model_fields") and "file_id" in BulkDraftImage.model_fields:
                        images.append(BulkDraftImage(filename=fname, file_id=uploaded_file.id))
                    else:
                        data_url = await file_svc.data_url(uploaded_file)
                        images.append(BulkDraftImage(filename=fname, data_url=data_url))

            # 5b: image_urls column (remote URLs) — kept as-is for compatibility;
            # URL-based images were handled by the old client-side path.  We
            # do not download remote URLs server-side in this wave; just pop
            # the column so it doesn't clutter the canonical dict.
            canonical.pop("image_urls", None)

            # ---- Construct BulkDraftRow ---------------------------------- #
            draft_rows.append(
                BulkDraftRow(
                    name=canonical.get("name"),
                    category_title=canonical.get("category_title"),
                    brief=canonical.get("brief"),
                    description=canonical.get("description"),
                    keywords=canonical.get("keywords"),
                    sku=canonical.get("sku"),
                    barcode=canonical.get("barcode"),
                    primary_price=canonical.get("primary_price"),
                    stock=canonical.get("stock"),
                    weight_g=canonical.get("weight_g"),
                    package_weight_g=canonical.get("package_weight_g"),
                    preparation_days=canonical.get("preparation_days"),
                    unit_quantity=canonical.get("unit_quantity"),
                    unit_type_title=canonical.get("unit_type_title"),
                    is_wholesale=canonical.get("is_wholesale"),
                    images=images,
                )
            )

        if not draft_rows:
            raise HTTPException(
                status_code=400,
                detail=(
                    "هیچ ردیف معتبری در شیت یافت نشد. "
                    "مطمئن شوید ستون‌های اجباری پر هستند."
                ),
            )

        # ------------------------------------------------------------------ #
        # 6. Delegate to ProductService.save_drafts
        # ------------------------------------------------------------------ #
        from app.services.product_service import ProductService

        self.logger.info(
            "save_drafts_from_files calling save_drafts user_id=%s rows=%d",
            user.id, len(draft_rows),
        )
        created = await ProductService(self.session).save_drafts(
            user,
            BulkSaveDraftsRequest(rows=draft_rows),
        )
        self.logger.info(
            "save_drafts_from_files complete user_id=%s created=%d",
            user.id, len(created),
        )
        return created
