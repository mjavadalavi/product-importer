from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.services.bulk_import.columns import COLUMNS, ColumnSpec

# Three sample rows so the user sees a complete fill pattern.
SAMPLE_ROWS: list[dict[str, str]] = [
    {
        "name": "عسل طبیعی کوهی ۵۰۰ گرمی",
        "category_title": "عسل",
        "primary_price": "285000",
        "stock": "12",
        "weight_g": "500",
        "package_weight_g": "560",
        "preparation_days": "3",
        "brief": "عسل خام و طبیعی از کوه‌های زاگرس",
        "description": "عسل کاملاً طبیعی، بدون افزودنی و حرارت‌ندیده. مناسب مصرف روزانه.",
        "keywords": "عسل، طبیعی، کوهی، ارگانیک",
        "sku": "HONEY-500-A1",
        "barcode": "1234567890123",
        "unit_quantity": "500",
        "unit_type_title": "گرم",
        "is_wholesale": "false",
        "image_urls": "https://example.com/honey-1.jpg, https://example.com/honey-2.jpg",
        "image_filenames": "",
    },
    {
        "name": "شال زنانه نخی طرح‌دار",
        "category_title": "شال و روسری",
        "primary_price": "180000",
        "stock": "8",
        "weight_g": "120",
        "package_weight_g": "150",
        "preparation_days": "5",
        "brief": "شال نخی سبک با طرح گل",
        "description": "شال زنانه از جنس نخ مرغوب، مناسب چهارفصل.",
        "keywords": "شال، نخی، زنانه، گل‌دار",
        "sku": "SCARF-NX-12",
        "barcode": "",
        "unit_quantity": "1",
        "unit_type_title": "عددی",
        "is_wholesale": "false",
        "image_urls": "",
        "image_filenames": "scarf-1.jpg, scarf-2.jpg",
    },
    {
        "name": "کاسه سفالی دست‌ساز سایز متوسط",
        "category_title": "ظروف سفالی",
        "primary_price": "95000",
        "stock": "25",
        "weight_g": "350",
        "package_weight_g": "420",
        "preparation_days": "2",
        "brief": "کاسه سفالی لعاب‌دار، دست‌ساز",
        "description": "کاسه‌ی سفالی با لعاب آبی، مناسب سرو سوپ یا سالاد.",
        "keywords": "سفال، دست‌ساز، کاسه، صنایع‌دستی",
        "sku": "POT-M-04",
        "barcode": "",
        "unit_quantity": "1",
        "unit_type_title": "عددی",
        "is_wholesale": "false",
        "image_urls": "",
        "image_filenames": "pot-m-1.jpg",
    },
]

HEADER_FILL = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFFFE9B0", end_color="FFFFE9B0", fill_type="solid")
GUIDE_FILL = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")


def _write_data_sheet(ws: Worksheet) -> None:
    ws.title = "محصولات"
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, size=11)
    required_font = Font(bold=True, size=11, color="FF8A4500")

    for idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=col.title)
        cell.font = required_font if col.required else header_font
        cell.fill = REQUIRED_FILL if col.required else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # widen by header text + a little padding
        ws.column_dimensions[get_column_letter(idx)].width = max(18, len(col.title) + 4)

    ws.row_dimensions[1].height = 32

    for r_idx, row in enumerate(SAMPLE_ROWS, start=2):
        for c_idx, col in enumerate(COLUMNS, start=1):
            value = row.get(col.key, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

    # leave 50 empty rows for the user to fill in
    ws.freeze_panes = "A2"

    # is_wholesale dropdown
    is_wholesale_col = next(
        (i + 1 for i, c in enumerate(COLUMNS) if c.key == "is_wholesale"), None
    )
    if is_wholesale_col:
        letter = get_column_letter(is_wholesale_col)
        dv = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
        dv.error = "فقط مقدار true یا false مجاز است."
        dv.errorTitle = "مقدار نامعتبر"
        dv.prompt = "true یا false"
        dv.promptTitle = "عمده‌فروشی"
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}1000")


def _write_guide_sheet(ws: Worksheet) -> None:
    ws.title = "راهنما"
    ws.sheet_view.rightToLeft = True

    title_cell = ws.cell(row=1, column=1, value="راهنمای پر کردن ستون‌ها")
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 24

    headers = ["ستون", "اجباری؟", "توضیح", "نمونه"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = GUIDE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [22, 12, 60, 38]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for offset, col in enumerate(COLUMNS):
        r = 4 + offset
        ws.cell(row=r, column=1, value=col.title)
        ws.cell(row=r, column=2, value="بله" if col.required else "—")
        ws.cell(row=r, column=3, value=col.description)
        ws.cell(row=r, column=4, value=col.example)
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 28

    # Notes block
    notes_start = 4 + len(COLUMNS) + 2
    note_lines = [
        "نکته‌ها:",
        "- ستون «دسته‌بندی» را با نام دقیق فارسی پر کنید؛ سرور بر اساس آن دسته‌ی باسلام را پیدا می‌کند.",
        "- قیمت‌ها به تومان است، بدون کاما یا واحد.",
        "- وزن‌ها به گرم است.",
        "- اگر عکس‌ها را با ZIP می‌فرستی، فایل اکسل را در ریشه و عکس‌ها را داخل پوشه images/ بگذار و در ستون «نام فایل عکس‌ها در ZIP» نام دقیق فایل‌ها را وارد کن.",
        "- ستون‌های اختیاری را اگر نمی‌دانی خالی بگذار؛ بعد از وارد کردن از داشبورد می‌توانی ویرایش کنی.",
    ]
    for i, line in enumerate(note_lines):
        cell = ws.cell(row=notes_start + i, column=1, value=line)
        cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        if i == 0:
            cell.font = Font(bold=True)
        ws.merge_cells(
            start_row=notes_start + i,
            start_column=1,
            end_row=notes_start + i,
            end_column=4,
        )


def build_template_xlsx() -> bytes:
    """Build the Persian Basalam product-import template xlsx as bytes."""
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    data_ws = wb.create_sheet("محصولات")
    _write_data_sheet(data_ws)
    guide_ws = wb.create_sheet("راهنما")
    _write_guide_sheet(guide_ws)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
